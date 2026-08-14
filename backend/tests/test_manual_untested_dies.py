"""人工设为未测试的 Die：持久化、统计排除和恢复。"""

from __future__ import annotations

import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app import store
from app.data_source import _aggregate_shots
from app.export_excel import build_judge_workbook
from app.judge import summarize


class ManualUntestedStoreTests(unittest.TestCase):
    def setUp(self) -> None:
        self._old_data_dir = store.DATA_DIR
        self._old_db_path = store.DB_PATH
        self._tmp = tempfile.TemporaryDirectory(ignore_cleanup_errors=True)
        store.DATA_DIR = Path(self._tmp.name)
        store.DB_PATH = store.DATA_DIR / "test.db"
        store.init_db()

    def tearDown(self) -> None:
        store.DATA_DIR = self._old_data_dir
        store.DB_PATH = self._old_db_path
        self._tmp.cleanup()

    def test_manual_untested_state_can_be_saved_and_restored_per_wafer(self) -> None:
        store.set_die_manual_untested("WAFER-A", "die-1", True)
        store.set_die_manual_untested("WAFER-A", "die-2", True)
        store.set_die_manual_untested("WAFER-B", "die-1", True)

        self.assertEqual(store.list_manual_untested_die_ids("WAFER-A"), {"die-1", "die-2"})
        self.assertEqual(store.list_manual_untested_die_ids("WAFER-B"), {"die-1"})

        store.set_die_manual_untested("WAFER-A", "die-1", False)
        self.assertEqual(store.list_manual_untested_die_ids("WAFER-A"), {"die-2"})

    def test_manual_untested_state_survives_repeated_refreshes_and_db_reinitialization(self) -> None:
        store.set_die_manual_untested("WAFER-A", "die-stable", True)

        self.assertEqual(store.list_manual_untested_die_ids("WAFER-A"), {"die-stable"})
        self.assertEqual(store.list_manual_untested_die_ids("WAFER-A"), {"die-stable"})
        store.init_db()
        self.assertEqual(store.list_manual_untested_die_ids("WAFER-A"), {"die-stable"})

        store.set_die_manual_untested("WAFER-A", "die-stable", False)
        self.assertEqual(store.list_manual_untested_die_ids("WAFER-A"), set())


class ManualUntestedSummaryTests(unittest.TestCase):
    def test_excluded_die_stays_in_shot_but_is_removed_from_all_counts(self) -> None:
        dies = [
            {
                "id": "pass-die",
                "shot": "51",
                "serial": "0101",
                "pass": True,
                "manual_untested": False,
                "param_rows": [{"key": "Loss", "name": "Loss", "pass": True}],
            },
            {
                "id": "excluded-fail-die",
                "shot": "51",
                "serial": "0102",
                "pass": False,
                "manual_untested": True,
                "param_rows": [{"key": "Loss", "name": "Loss", "pass": False}],
            },
        ]

        stats = summarize(dies, [{"name": "Loss", "enabled": True}])
        shot = _aggregate_shots(dies)[0]

        self.assertEqual(stats["total"], 1)
        self.assertEqual(stats["pass_count"], 1)
        self.assertEqual(stats["fail_count"], 0)
        self.assertEqual(stats["fail_rate_details"][0]["fail_count"], 0)
        self.assertEqual(shot["die_count"], 1)
        self.assertEqual(shot["pass_count"], 1)
        self.assertEqual(shot["fail_count"], 0)
        self.assertEqual(len(shot["dies"]), 2)
        excluded = next(d for d in shot["dies"] if d["id"] == "excluded-fail-die")
        self.assertTrue(excluded["manual_untested"])

    def test_export_labels_excluded_die_untested_and_omits_it_from_fail_list(self) -> None:
        payload = {
            "wafer": "WAFER-A",
            "stats": {"total": 0, "pass_count": 0, "fail_count": 0, "yield": 0, "fail_rate_details": []},
            "dies": [
                {
                    "shot": "51",
                    "serial": "0101",
                    "pass": False,
                    "manual_untested": True,
                    "param_rows": [{"name": "Loss", "pass": False}],
                }
            ],
        }

        workbook = load_workbook(BytesIO(build_judge_workbook(payload)))
        self.assertEqual(workbook["全部Die"]["G2"].value, "未测试")
        self.assertEqual(workbook["不良品清单"].max_row, 1)


if __name__ == "__main__":
    unittest.main()
