from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook


def build_judge_workbook(payload: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "判定汇总"
    stats = payload.get("stats") or {}
    ws.append(["晶圆编号", payload.get("wafer")])
    ws.append(["总芯片数", stats.get("total")])
    ws.append(["良品数", stats.get("pass_count")])
    ws.append(["不良品数", stats.get("fail_count")])
    ws.append(["良率(%)", stats.get("yield")])
    ws.append([])
    ws.append(["参数", "不良数", "不良率(%)"])
    for row in stats.get("fail_rate_details") or []:
        ws.append([row.get("name"), row.get("fail_count"), row.get("fail_rate")])

    ws2 = wb.create_sheet("全部Die")
    ws2.append(["Shot", "SN", "流水号", "X", "Y", "CreateTime", "判定", "失败参数"])
    for die in payload.get("dies") or []:
        fails = [
            r["name"]
            for r in die.get("param_rows") or []
            if r.get("pass") is False and r.get("name") != "总体评价" and not r.get("is_overall")
        ]
        ws2.append(
            [
                die.get("shot"),
                die.get("sn"),
                die.get("serial"),
                die.get("x"),
                die.get("y"),
                die.get("create_time"),
                "Pass" if die.get("pass") else "Fail",
                ",".join(fails),
            ]
        )

    ws3 = wb.create_sheet("不良品清单")
    ws3.append(["Shot", "SN", "流水号", "X", "Y", "参数", "单位", "实测值", "LSL", "USL"])
    for die in payload.get("dies") or []:
        if die.get("pass"):
            continue
        for r in die.get("param_rows") or []:
            if r.get("pass") is False and r.get("name") != "总体评价" and not r.get("is_overall"):
                ws3.append(
                    [
                        die.get("shot"),
                        die.get("sn"),
                        die.get("serial"),
                        die.get("x"),
                        die.get("y"),
                        r.get("name"),
                        r.get("unit"),
                        r.get("value"),
                        r.get("lsl"),
                        r.get("usl"),
                    ]
                )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
